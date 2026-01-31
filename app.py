from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
FILE_NAME = "/tmp/client_data.xlsx"


# Create Excel file if not exists
if not os.path.exists(FILE_NAME):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Clients"
    ws1.append(["Client Name", "Company", "Email", "Phone"])

    ws2 = wb.create_sheet("Meetings")
    ws2.append(["Client Name", "Meeting Date", "Summary", "Next Steps", "Actions"])

    wb.save(FILE_NAME)

@app.route('/')
def home():
    return render_template("index.html")

@app.route('/add_client', methods=['POST'])
def add_client():
    wb = load_workbook(FILE_NAME)
    ws = wb["Clients"]
    ws.append([
        request.form['client_name'],
        request.form['company'],
        request.form['email'],
        request.form['phone']
    ])
    wb.save(FILE_NAME)
    return "Client Added Successfully!"

@app.route('/add_meeting', methods=['POST'])
def add_meeting():
    wb = load_workbook(FILE_NAME)
    ws = wb["Meetings"]
    ws.append([
        request.form['client_name'],
        request.form['meeting_date'],
        request.form['summary'],
        request.form['next_steps'],
        request.form['actions']
    ])
    wb.save(FILE_NAME)
    return "Meeting Logged Successfully!"

if __name__ == '__main__':
    import os
port = int(os.environ.get("PORT", 5000))
app.run(host="0.0.0.0", port=port)

